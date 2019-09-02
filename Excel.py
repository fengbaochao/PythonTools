# coding:utf-8

"""
Excel 工具类
@Excel
@Author:Feng
@Date:2019-08-30
@QQ:960059842
"""

import xlrd, xlwt
from openpyxl import Workbook
from Tools.Path import Path

class Excel:
    def read(self, path, sheetIndex=0):
        data = xlrd.open_workbook(path)
        table = data.sheets()[sheetIndex]  # 用索引取第一个sheet

        nrows = table.nrows  # 行数
        ncols = table.ncols  # 列数
        data = []

        for i in range(0, nrows):
            rowValues = table.row_values(i)  # 某一行数据
            data.append(rowValues)

        return data

    def __writeExcel2007(self, filePath, sheetName, cols, data):
        workBook = Workbook()  # 创建文件对象
        sheet = workBook.active  # 获取第一个sheet
        sheet.title = sheetName

        # 写入标题
        rowIndex = 1
        for i in range(len(cols)):
            sheet.cell(row=rowIndex, column=i + 1, value=cols[i])

        # 取出data中的每一个元组存到表格的每一行
        for values in data:
            rowIndex += 1
            for i in range(len(values)):
                sheet.cell(row=rowIndex, column=i + 1, value=values[i])

        workBook.save(filePath)  # 保存excel

    def __writeExcel2003(self, filePath, sheetName, cols, data):
        workBook = xlwt.Workbook()  # 创建excel对象
        sheet = workBook.add_sheet(sheetName)  # 添加一个表

        # 写入标题
        rowIndex = 0
        for i in range(len(cols)):
            sheet.write(rowIndex, i, cols[i])

        # 取出data中的每一个元组存到表格的每一行
        for values in data:
            rowIndex += 1
            for i in range(len(values)):
                sheet.write(rowIndex, i, values[i])

        workBook.save(filePath)  # 保存excel

    def write(self, filePath, cols, data):
        path = Path()
        ext = path.getExtension(filePath)
        sheetName = path.getFileNameWithoutExtension(filePath)
        if ext == ".xls":
            self.__writeExcel2003(filePath, sheetName, cols, data)
        else:
            self.__writeExcel2007(filePath, sheetName, cols, data)


if __name__ == "__main__":
    # path = r'C:\Users\FBC\Desktop\题库\011建筑电工.xlsx'
    # path = r'C:\Users\FBC\Desktop\题库\011建筑电工.xls'
    excel = Excel()
    # print(excel.read(path))

    # path = r"C:\Users\FBC\Desktop\题库\001.xls"
    path = r'C:\Users\FBC\Desktop\题库\001.xlsx'
    cols = ["列1", "列2", "列3", "列4"]
    data = []
    for i in range(5):
        data.append(["张三" + str(i), "李四" + str(i), "王五" + str(i), "赵六" + str(i)])
    excel.write(path, cols, data)
